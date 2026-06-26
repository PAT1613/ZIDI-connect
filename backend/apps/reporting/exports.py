from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def rows_to_xlsx(title: str, rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, value in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(12, width + 2))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_pdf(title: str, rows: list[list]) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    flow = [
        Paragraph(f"<b>ZIDI Connect — {title}</b>", styles["Title"]),
        Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Italic"]),
        Spacer(1, 6 * mm),
    ]
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    flow.append(table)
    doc.build(flow)
    return buf.getvalue()
